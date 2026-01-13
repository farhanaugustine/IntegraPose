from __future__ import annotations

import argparse
import csv
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

import cv2


@dataclass
class Segment:
    start_time: float
    end_time: float
    start_frame: int
    end_frame: int
    label: str
    confidence: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Overlay CSV annotations on the source video to visually audit the "
            "behavior predictions."
        )
    )
    parser.add_argument("--video_path", type=Path, required=True)
    parser.add_argument(
        "--annotations_csv",
        type=Path,
        default=None,
        help="Merged CSV produced by inference (used for legacy panel overlay).",
    )
    parser.add_argument(
        "--metrics_xlsx",
        type=Path,
        default=None,
        help="Metrics workbook produced by inference (--output_xlsx). If provided, draws YOLO-style bbox labels instead of the top panel.",
    )
    parser.add_argument(
        "--output_video",
        type=Path,
        default=None,
        help="Optional path to save an annotated video (e.g., review.mp4).",
    )
    parser.add_argument(
        "--no_display",
        action="store_true",
        help="Skip the interactive preview window (useful when only saving video).",
    )
    parser.add_argument(
        "--font_scale",
        type=float,
        default=0.7,
        help="OpenCV font scale for overlay text.",
    )
    parser.add_argument(
        "--thickness",
        type=int,
        default=2,
        help="Text/border thickness for overlays.",
    )
    parser.add_argument(
        "--label_bg_alpha",
        type=float,
        default=1.0,
        help=(
            "Opacity (0-1) for the black text background rectangles. "
            "Lower values reduce occlusion during review."
        ),
    )
    parser.add_argument(
        "--wait_ms",
        type=int,
        default=1,
        help="Delay (ms) for cv2.waitKey. Increase to slow playback.",
    )
    return parser.parse_args()


def _clamp_alpha(alpha: float | None, *, default: float = 1.0) -> float:
    if alpha is None:
        return float(default)
    try:
        alpha_f = float(alpha)
    except Exception:
        return float(default)
    if not math.isfinite(alpha_f):
        return float(default)
    return max(0.0, min(1.0, alpha_f))


def _fill_rect_alpha(
    image,
    x1: int,
    y1: int,
    x2: int,
    y2: int,
    color: tuple[int, int, int],
    alpha: float,
) -> None:
    alpha = _clamp_alpha(alpha, default=1.0)
    if alpha <= 0.0:
        return

    h, w = image.shape[:2]
    x1 = int(max(0, min(w - 1, x1)))
    y1 = int(max(0, min(h - 1, y1)))
    x2 = int(max(0, min(w - 1, x2)))
    y2 = int(max(0, min(h - 1, y2)))
    if x2 < x1:
        x1, x2 = x2, x1
    if y2 < y1:
        y1, y2 = y2, y1

    if alpha >= 1.0:
        cv2.rectangle(image, (x1, y1), (x2, y2), color, thickness=cv2.FILLED)
        return

    roi = image[y1 : y2 + 1, x1 : x2 + 1]
    if roi.size == 0:
        return
    layer = roi.copy()
    layer[:] = color
    cv2.addWeighted(layer, alpha, roi, 1.0 - alpha, 0.0, dst=roi)


def load_segments(csv_path: Path) -> List[Segment]:
    segments: List[Segment] = []
    with csv_path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                segments.append(
                    Segment(
                        start_time=float(row["start_time"]),
                        end_time=float(row["end_time"]),
                        start_frame=int(row["start_frame"]),
                        end_frame=int(row["end_frame"]),
                        label=row["label"],
                        confidence=float(row["confidence"]),
                    )
                )
            except Exception as exc:  # pragma: no cover
                print(f"Warning: skipping malformed row {row} ({exc})")
    segments.sort(key=lambda seg: seg.start_frame)
    return segments


def annotate_frame(
    frame,
    segment: Optional[Segment],
    frame_idx: int,
    fps: float,
    total_frames: int,
    font_scale: float,
    thickness: int,
    *,
    label_bg_alpha: float = 1.0,
):
    overlay = frame.copy()
    h, w = frame.shape[:2]
    panel_height = 90
    _fill_rect_alpha(
        overlay,
        0,
        0,
        int(w - 1),
        int(panel_height),
        (0, 0, 0),
        float(label_bg_alpha),
    )

    if segment:
        label = segment.label
        conf = segment.confidence
        time_text = f"{segment.start_time:.2f}s - {segment.end_time:.2f}s"
    else:
        label = "No Annotation"
        conf = 0.0
        time_text = "N/A"

    lines = [
        f"Frame {frame_idx}  Time {frame_idx / max(fps, 1e-3):.2f}s",
        f"Label: {label}  Confidence: {conf:.2f}",
        f"Segment: {time_text}",
    ]

    y = 30
    for line in lines:
        cv2.putText(
            overlay,
            line,
            (10, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            (255, 255, 255),
            thickness,
            lineType=cv2.LINE_AA,
        )
        y += int(30 * font_scale)

    progress = frame_idx / max(total_frames - 1, 1)
    bar_y = panel_height - 10
    cv2.rectangle(
        overlay,
        (0, bar_y),
        (int(w * progress), bar_y + 5),
        (0, 255, 0) if segment else (0, 128, 255),
        thickness=cv2.FILLED,
    )
    return overlay


def annotate_frame_bbox(
    frame,
    bbox_xyxy: tuple[int, int, int, int] | None,
    label: str,
    confidence: float | None,
    yolo_status: str | None,
    font_scale: float,
    thickness: int,
    *,
    keypoints: list[tuple[float, float, float | None]] | None = None,
    draw_keypoints: bool = True,
    keypoint_radius: int = 3,
    show_keypoint_indices: bool = False,
    label_bg_alpha: float = 1.0,
):
    overlay = frame.copy()
    h, w = overlay.shape[:2]
    bg_alpha = _clamp_alpha(float(label_bg_alpha), default=1.0)
    if bbox_xyxy is not None:
        x1, y1, x2, y2 = bbox_xyxy
        cv2.rectangle(overlay, (x1, y1), (x2, y2), (0, 255, 0), thickness)
        text = label
        if confidence is not None:
            text = f"{label} {confidence:.2f}"
        if yolo_status and yolo_status != "detected":
            text = f"{text} [{yolo_status}]"
        (tw, th), baseline = cv2.getTextSize(
            text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness
        )
        box_w = int(tw + 6)
        box_h = int(th + baseline + 6)
        tx = int(max(0, min(int(x1), int(w - box_w))))
        ty_above = int(y1) - box_h - 2
        if ty_above >= 0:
            ty = int(ty_above)
        else:
            ty_below = int(y2) + 2
            if ty_below + box_h <= h:
                ty = int(ty_below)
            else:
                ty = int(max(0, ty_above))
        _fill_rect_alpha(
            overlay,
            tx,
            ty,
            tx + box_w,
            ty + box_h,
            (0, 0, 0),
            bg_alpha,
        )
        cv2.putText(
            overlay,
            text,
            (tx + 3, ty + th + 3),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            (255, 255, 255),
            thickness,
            lineType=cv2.LINE_AA,
        )
    else:
        text = label
        if yolo_status == "missed":
            text = f"{label} [yolo_missed]"
        (tw, th), baseline = cv2.getTextSize(
            text, cv2.FONT_HERSHEY_SIMPLEX, font_scale, thickness
        )
        tx = 10
        ty = max(0, 30 - int(th + baseline) - 4)
        box_w = int(tw + 6)
        box_h = int(th + baseline + 6)
        _fill_rect_alpha(
            overlay,
            tx,
            ty,
            min(int(w - 1), tx + box_w),
            min(int(h - 1), ty + box_h),
            (0, 0, 0),
            bg_alpha,
        )
        cv2.putText(
            overlay,
            text,
            (tx + 3, ty + th + 3),
            cv2.FONT_HERSHEY_SIMPLEX,
            font_scale,
            (255, 255, 255),
            thickness,
            lineType=cv2.LINE_AA,
        )
    if draw_keypoints and keypoints:
        r = max(1, int(keypoint_radius))
        font_small = max(0.4, float(font_scale) * 0.6)
        for kpt_idx, (x, y, c) in enumerate(keypoints):
            try:
                px = int(round(float(x)))
                py = int(round(float(y)))
            except Exception:
                continue
            color = (255, 255, 0)
            if c is not None:
                try:
                    c_f = max(0.0, min(1.0, float(c)))
                    color = (0, int(255.0 * c_f), int(255.0 * (1.0 - c_f)))
                except Exception:
                    color = (255, 255, 0)
            cv2.circle(overlay, (px, py), r, color, thickness=cv2.FILLED)
            if show_keypoint_indices:
                cv2.putText(
                    overlay,
                    str(kpt_idx),
                    (px + r + 1, py - r - 1),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    font_small,
                    (255, 255, 255),
                    max(1, int(thickness) - 1),
                    lineType=cv2.LINE_AA,
                )
    return overlay


def iter_metrics_per_frame(metrics_xlsx: Path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - optional dependency
        raise ImportError(
            "openpyxl is required to read metrics_xlsx. Install it with: pip install openpyxl"
        ) from exc

    wb = load_workbook(metrics_xlsx, read_only=True, data_only=True)
    sheets = [ws for ws in wb.worksheets if ws.title.startswith("per_frame")]
    if not sheets:
        raise ValueError("No per_frame sheet found in metrics workbook.")

    for ws in sheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        idx = {str(name): i for i, name in enumerate(header) if name is not None}
        kpt_tokens: list[str] = []
        kpt_x_re = re.compile(r"^(kpt_\d+)_x$")
        for name in idx:
            m = kpt_x_re.match(name)
            if m:
                kpt_tokens.append(str(m.group(1)))
        kpt_tokens = sorted(
            set(kpt_tokens), key=lambda token: int(token.split("_", 1)[1])
        )
        for row in rows:
            if not row:
                continue
            try:
                frame_idx = int(row[idx["frame_idx"]])
            except Exception:
                continue
            def _get(name: str):
                j = idx.get(name)
                if j is None or j >= len(row):
                    return None
                return row[j]

            x1 = _get("bbox_x1")
            y1 = _get("bbox_y1")
            x2 = _get("bbox_x2")
            y2 = _get("bbox_y2")
            bbox = None
            if x1 is not None and y1 is not None and x2 is not None and y2 is not None:
                try:
                    bbox = (int(x1), int(y1), int(x2), int(y2))
                except Exception:
                    bbox = None

            label = _get("behavior_label")
            conf = _get("behavior_conf")
            status = _get("yolo_status")
            keypoints = None
            if kpt_tokens:
                pts: list[tuple[float, float, float | None]] = []
                for token in kpt_tokens:
                    x = _get(f"{token}_x")
                    y = _get(f"{token}_y")
                    c = _get(f"{token}_conf")
                    if x is None or y is None:
                        continue
                    try:
                        fx = float(x)
                        fy = float(y)
                    except Exception:
                        continue
                    fc = None
                    if c is not None:
                        try:
                            fc = float(c)
                        except Exception:
                            fc = None
                    pts.append((fx, fy, fc))
                if pts:
                    keypoints = pts
            yield {
                "frame_idx": frame_idx,
                "bbox_xyxy": bbox,
                "label": str(label) if label is not None else "Unknown",
                "confidence": float(conf) if conf is not None else None,
                "yolo_status": str(status) if status is not None else None,
                "keypoints": keypoints,
            }


def main():
    args = parse_args()
    video_path = args.video_path
    csv_path = args.annotations_csv
    metrics_xlsx = args.metrics_xlsx

    if not video_path.exists():
        raise FileNotFoundError(f"Video not found: {video_path}")
    if metrics_xlsx is None and csv_path is None:
        raise ValueError("Provide either --metrics_xlsx or --annotations_csv.")
    if metrics_xlsx is not None and not metrics_xlsx.exists():
        raise FileNotFoundError(f"Metrics XLSX not found: {metrics_xlsx}")
    if metrics_xlsx is None:
        if csv_path is None or not csv_path.exists():
            raise FileNotFoundError(f"CSV not found: {csv_path}")

    segments: list[Segment] = []
    metrics_iter = None
    current_metric = None
    if metrics_xlsx is not None:
        metrics_iter = iter_metrics_per_frame(metrics_xlsx)
        current_metric = next(metrics_iter, None)
    else:
        segments = load_segments(csv_path)  # type: ignore[arg-type]
        if not segments:
            print(
                "No annotations were loaded. Check that the CSV has the expected header."
            )

    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        raise RuntimeError(f"Unable to open video: {video_path}")

    fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT) or 0)
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    writer = None
    if args.output_video:
        fourcc = cv2.VideoWriter_fourcc(*"mp4v")
        writer = cv2.VideoWriter(
            str(args.output_video), fourcc, fps, (width, height)
        )
        if not writer.isOpened():
            raise RuntimeError(f"Unable to open video writer: {args.output_video}")

    current_idx = 0
    frame_idx = 0
    window_name = "Behavior Review"
    if not args.no_display:
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        if metrics_iter is not None:
            while current_metric is not None and int(current_metric["frame_idx"]) < frame_idx:
                current_metric = next(metrics_iter, None)
            if current_metric is not None and int(current_metric["frame_idx"]) == frame_idx:
                annotated = annotate_frame_bbox(
                    frame,
                    current_metric.get("bbox_xyxy"),
                    str(current_metric.get("label", "Unknown")),
                    current_metric.get("confidence"),
                    current_metric.get("yolo_status"),
                    args.font_scale,
                    args.thickness,
                    keypoints=current_metric.get("keypoints"),
                    label_bg_alpha=float(args.label_bg_alpha),
                )
            else:
                annotated = frame
        else:
            while (
                current_idx < len(segments)
                and frame_idx > segments[current_idx].end_frame
            ):
                current_idx += 1

            if (
                current_idx < len(segments)
                and segments[current_idx].start_frame
                <= frame_idx
                <= segments[current_idx].end_frame
            ):
                active = segments[current_idx]
            else:
                active = None

            annotated = annotate_frame(
                frame,
                active,
                frame_idx,
                fps,
                total_frames,
                args.font_scale,
                args.thickness,
                label_bg_alpha=float(args.label_bg_alpha),
            )

        if writer:
            writer.write(annotated)

        if not args.no_display:
            cv2.imshow(window_name, annotated)
            key = cv2.waitKey(args.wait_ms) & 0xFF
            if key == ord("q"):
                break

        frame_idx += 1

    cap.release()
    if writer:
        writer.release()
    if not args.no_display:
        cv2.destroyAllWindows()


if __name__ == "__main__":
    main()
