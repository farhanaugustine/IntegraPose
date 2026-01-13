from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import numpy as np

EXCEL_MAX_ROWS = 1_048_576
MAX_TOP_K = 5


def clamp_top_k(top_k: int, num_classes: int) -> int:
    top_k = int(top_k)
    if top_k <= 0:
        top_k = 1
    top_k = min(top_k, MAX_TOP_K)
    top_k = min(top_k, int(num_classes))
    return max(top_k, 1)


def compute_centers_xy(
    bboxes_xyxy: Sequence[tuple[int, int, int, int] | None],
) -> np.ndarray:
    centers = np.full((len(bboxes_xyxy), 2), np.nan, dtype=np.float32)
    for i, bbox in enumerate(bboxes_xyxy):
        if bbox is None:
            continue
        x1, y1, x2, y2 = bbox
        centers[i, 0] = (float(x1) + float(x2)) / 2.0
        centers[i, 1] = (float(y1) + float(y2)) / 2.0
    return centers


def smooth_centers_xy(
    centers_xy: np.ndarray,
    method: str,
    *,
    ema_alpha: float = 0.2,
    interp_max_gap: int = 5,
    kalman_process_var: float = 0.05,
    kalman_measure_var: float = 25.0,
) -> np.ndarray:
    method = str(method or "none").strip().lower()
    if method in {"none", "raw", ""}:
        return centers_xy.astype(np.float32, copy=True)
    if method in {"ema", "exp", "exponential"}:
        return _smooth_ema(centers_xy, alpha=float(ema_alpha))
    if method in {"kalman", "kf"}:
        return _smooth_kalman(
            centers_xy,
            process_var=float(kalman_process_var),
            measure_var=float(kalman_measure_var),
        )
    if method in {"interp", "interpolate", "interpolation", "linear"}:
        return _smooth_interpolate(centers_xy, max_gap=int(interp_max_gap))
    raise ValueError(f"Unknown smoothing method: {method}")


def _smooth_ema(centers_xy: np.ndarray, *, alpha: float) -> np.ndarray:
    alpha = float(alpha)
    if not (0.0 < alpha <= 1.0):
        alpha = 0.2
    out = np.full_like(centers_xy, np.nan, dtype=np.float32)
    prev = None
    for i in range(len(centers_xy)):
        x, y = centers_xy[i]
        if np.isfinite(x) and np.isfinite(y):
            if prev is None:
                prev = np.array([x, y], dtype=np.float32)
            else:
                prev = alpha * np.array([x, y], dtype=np.float32) + (
                    (1.0 - alpha) * prev
                )
            out[i] = prev
        else:
            if prev is not None:
                out[i] = prev
    return out


@dataclass
class _Kalman2D:
    process_var: float
    measure_var: float

    def __post_init__(self) -> None:
        dt = 1.0
        self.F = np.array(
            [[1.0, 0.0, dt, 0.0], [0.0, 1.0, 0.0, dt], [0.0, 0.0, 1.0, 0.0], [0.0, 0.0, 0.0, 1.0]],
            dtype=np.float32,
        )
        self.H = np.array([[1.0, 0.0, 0.0, 0.0], [0.0, 1.0, 0.0, 0.0]], dtype=np.float32)
        self.Q = np.eye(4, dtype=np.float32) * float(self.process_var)
        self.R = np.eye(2, dtype=np.float32) * float(self.measure_var)
        self.P = np.eye(4, dtype=np.float32) * 1000.0
        self.x = np.zeros((4, 1), dtype=np.float32)
        self.initialized = False

    def predict(self) -> None:
        self.x = self.F @ self.x
        self.P = self.F @ self.P @ self.F.T + self.Q

    def update(self, z: np.ndarray) -> None:
        y = z - (self.H @ self.x)
        S = self.H @ self.P @ self.H.T + self.R
        K = self.P @ self.H.T @ np.linalg.inv(S)
        self.x = self.x + (K @ y)
        I = np.eye(self.P.shape[0], dtype=np.float32)
        self.P = (I - (K @ self.H)) @ self.P


def _smooth_kalman(
    centers_xy: np.ndarray, *, process_var: float, measure_var: float
) -> np.ndarray:
    out = np.full_like(centers_xy, np.nan, dtype=np.float32)
    kf = _Kalman2D(process_var=process_var, measure_var=measure_var)
    for i in range(len(centers_xy)):
        x, y = centers_xy[i]
        if not kf.initialized:
            if np.isfinite(x) and np.isfinite(y):
                kf.x[:2, 0] = np.array([x, y], dtype=np.float32)
                kf.x[2:, 0] = 0.0
                kf.initialized = True
                out[i] = np.array([x, y], dtype=np.float32)
            continue

        kf.predict()
        if np.isfinite(x) and np.isfinite(y):
            z = np.array([[x], [y]], dtype=np.float32)
            kf.update(z)
        out[i] = kf.x[:2, 0]
    return out


def _smooth_interpolate(centers_xy: np.ndarray, *, max_gap: int) -> np.ndarray:
    max_gap = int(max_gap)
    if max_gap <= 0:
        return centers_xy.astype(np.float32, copy=True)

    out = centers_xy.astype(np.float32, copy=True)
    valid = np.isfinite(out[:, 0]) & np.isfinite(out[:, 1])
    idx = np.flatnonzero(valid)
    if len(idx) == 0:
        return out

    first = int(idx[0])
    out[:first] = out[first]
    last = int(idx[-1])
    out[last + 1 :] = out[last]

    for left, right in zip(idx[:-1], idx[1:]):
        left = int(left)
        right = int(right)
        gap = right - left - 1
        if gap <= 0:
            continue
        if gap > max_gap:
            continue
        a = out[left]
        b = out[right]
        for k in range(1, gap + 1):
            t = k / float(gap + 1)
            out[left + k] = (1.0 - t) * a + t * b
    return out


def compute_kinematics(
    centers_xy: np.ndarray, fps: float
) -> dict[str, np.ndarray]:
    fps = float(fps) if fps and fps > 1e-6 else 30.0
    n = int(centers_xy.shape[0])
    step_dist = np.full(n, np.nan, dtype=np.float32)
    speed = np.full(n, np.nan, dtype=np.float32)
    accel = np.full(n, np.nan, dtype=np.float32)
    cum_dist = np.full(n, np.nan, dtype=np.float32)

    prev_xy = None
    prev_speed = None
    cum = 0.0
    for i in range(n):
        x, y = float(centers_xy[i, 0]), float(centers_xy[i, 1])
        if not (np.isfinite(x) and np.isfinite(y)):
            cum_dist[i] = float(cum)
            continue
        if prev_xy is None:
            prev_xy = (x, y)
            cum_dist[i] = float(cum)
            continue

        dx = x - prev_xy[0]
        dy = y - prev_xy[1]
        d = float(np.hypot(dx, dy))
        s = d * fps
        a = np.nan
        if prev_speed is not None and np.isfinite(prev_speed):
            a = (s - float(prev_speed)) * fps

        step_dist[i] = d
        speed[i] = s
        accel[i] = a
        cum += d
        cum_dist[i] = float(cum)
        prev_xy = (x, y)
        prev_speed = s

    return {
        "step_dist_px": step_dist,
        "speed_px_s": speed,
        "accel_px_s2": accel,
        "cum_dist_px": cum_dist,
    }


def windows_cover_count(
    total_frames: int, starts: np.ndarray, ends: np.ndarray
) -> np.ndarray:
    total_frames = int(total_frames)
    diff = np.zeros(total_frames + 1, dtype=np.int32)
    for s, e in zip(starts.tolist(), ends.tolist()):
        s = int(max(0, min(total_frames - 1, int(s))))
        e = int(max(0, min(total_frames - 1, int(e))))
        if e < s:
            s, e = e, s
        diff[s] += 1
        if e + 1 < len(diff):
            diff[e + 1] -= 1
    return np.cumsum(diff[:-1], dtype=np.int32)


def assign_windows_nearest_center(
    total_frames: int, centers: np.ndarray
) -> np.ndarray:
    total_frames = int(total_frames)
    if centers.size == 0:
        return np.full(total_frames, -1, dtype=np.int32)
    centers = centers.astype(np.int32, copy=False)
    n = int(len(centers))
    assignment = np.empty(total_frames, dtype=np.int32)

    bounds = np.empty(n + 1, dtype=np.int32)
    bounds[0] = 0
    for i in range(n - 1):
        bounds[i + 1] = int((int(centers[i]) + int(centers[i + 1])) // 2)
    bounds[n] = total_frames - 1

    start = 0
    for i in range(n):
        end = int(bounds[i + 1])
        if end < start:
            continue
        assignment[start : end + 1] = i
        start = end + 1
        if start >= total_frames:
            break

    if start < total_frames:
        assignment[start:] = n - 1
    return assignment


def aggregate_window_probs_average(
    total_frames: int,
    starts: np.ndarray,
    ends: np.ndarray,
    probs: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    total_frames = int(total_frames)
    probs = np.asarray(probs, dtype=np.float32)
    if probs.ndim != 2:
        raise ValueError("probs must be [num_windows, num_classes]")
    num_classes = int(probs.shape[1])
    sum_diff = np.zeros((total_frames + 1, num_classes), dtype=np.float32)
    count_diff = np.zeros(total_frames + 1, dtype=np.int32)
    for (s, e), p in zip(zip(starts.tolist(), ends.tolist()), probs):
        s = int(max(0, min(total_frames - 1, int(s))))
        e = int(max(0, min(total_frames - 1, int(e))))
        if e < s:
            s, e = e, s
        sum_diff[s] += p
        if e + 1 < len(sum_diff):
            sum_diff[e + 1] -= p
        count_diff[s] += 1
        if e + 1 < len(count_diff):
            count_diff[e + 1] -= 1
    sum_pf = np.cumsum(sum_diff[:-1], axis=0, dtype=np.float32)
    count_pf = np.cumsum(count_diff[:-1], dtype=np.int32)
    out = np.full_like(sum_pf, np.nan, dtype=np.float32)
    valid = count_pf > 0
    out[valid] = sum_pf[valid] / count_pf[valid, None].astype(np.float32)
    return out, count_pf


def topk_from_probs(
    probs: np.ndarray, top_k: int
) -> tuple[np.ndarray, np.ndarray]:
    probs = np.asarray(probs, dtype=np.float32)
    if probs.ndim != 2:
        raise ValueError("probs must be [N, C]")
    n, c = probs.shape
    k = clamp_top_k(int(top_k), int(c))
    idx = np.argpartition(probs, kth=c - k, axis=1)[:, -k:]
    vals = np.take_along_axis(probs, idx, axis=1)
    order = np.argsort(vals, axis=1)[:, ::-1]
    idx_sorted = np.take_along_axis(idx, order, axis=1)
    vals_sorted = np.take_along_axis(vals, order, axis=1)
    return idx_sorted.astype(np.int32), vals_sorted.astype(np.float32)


def bouts_from_frame_labels(
    labels: Sequence[str],
    fps: float,
    confidences: Sequence[float] | None = None,
    cum_dist_px: Sequence[float] | None = None,
) -> list[dict[str, object]]:
    fps = float(fps) if fps and fps > 1e-6 else 30.0
    n = len(labels)
    if n == 0:
        return []

    bouts: list[dict[str, object]] = []
    start = 0
    current = str(labels[0])
    for i in range(1, n):
        lab = str(labels[i])
        if lab == current:
            continue
        bouts.append(
            _bout_row(
                current,
                start,
                i - 1,
                fps,
                confidences=confidences,
                cum_dist_px=cum_dist_px,
            )
        )
        start = i
        current = lab
    bouts.append(
        _bout_row(
            current,
            start,
            n - 1,
            fps,
            confidences=confidences,
            cum_dist_px=cum_dist_px,
        )
    )
    return bouts


def _bout_row(
    label: str,
    start_frame: int,
    end_frame: int,
    fps: float,
    *,
    confidences: Sequence[float] | None,
    cum_dist_px: Sequence[float] | None,
) -> dict[str, object]:
    start_frame = int(start_frame)
    end_frame = int(end_frame)
    duration_frames = end_frame - start_frame + 1
    start_time = start_frame / fps
    end_time = end_frame / fps
    payload: dict[str, object] = {
        "label": label,
        "start_frame": start_frame,
        "end_frame": end_frame,
        "start_time": float(start_time),
        "end_time": float(end_time),
        "duration_s": float(duration_frames / fps),
    }
    if confidences is not None and len(confidences) > end_frame:
        window = confidences[start_frame : end_frame + 1]
        try:
            window_np = np.asarray(window, dtype=np.float32)
            if window_np.size == 0 or np.isnan(window_np).all():
                payload["mean_confidence"] = np.nan
            else:
                payload["mean_confidence"] = float(np.nanmean(window_np))
        except Exception:
            payload["mean_confidence"] = np.nan
    if cum_dist_px is not None and len(cum_dist_px) > end_frame:
        try:
            payload["distance_px"] = float(
                float(cum_dist_px[end_frame]) - float(cum_dist_px[start_frame])
            )
        except Exception:
            payload["distance_px"] = np.nan
    return payload


def write_metrics_xlsx(
    path: Path,
    *,
    per_frame_header: Sequence[str],
    per_frame_rows: Iterable[Sequence[object]],
    lstm_header: Sequence[str],
    lstm_rows: Iterable[Sequence[object]],
    bouts_header: Sequence[str],
    bouts_rows: Iterable[Sequence[object]],
    metadata: Mapping[str, object] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # pragma: no cover - optional dependency
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        ) from exc

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook(write_only=True)

    _write_split_sheet(
        wb,
        base_title="per_frame",
        header=per_frame_header,
        rows=per_frame_rows,
    )
    _write_simple_sheet(wb, "lstm_outputs", lstm_header, lstm_rows)
    _write_simple_sheet(wb, "bouts", bouts_header, bouts_rows)

    meta_items = list((metadata or {}).items())
    meta_rows = ([str(k), str(v)] for k, v in meta_items)
    _write_simple_sheet(wb, "metadata", ("key", "value"), meta_rows)

    wb.save(path)


def _write_simple_sheet(
    wb,
    title: str,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.append(list(header))
    for row in rows:
        ws.append(list(row))


def _write_split_sheet(
    wb,
    *,
    base_title: str,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> None:
    max_rows = EXCEL_MAX_ROWS - 1
    sheet_idx = 1
    ws = wb.create_sheet(title=base_title[:31])
    ws.append(list(header))
    rows_written = 0
    for row in rows:
        if rows_written >= max_rows:
            sheet_idx += 1
            title = f"{base_title}_{sheet_idx}"
            ws = wb.create_sheet(title=title[:31])
            ws.append(list(header))
            rows_written = 0
        ws.append(list(row))
        rows_written += 1
