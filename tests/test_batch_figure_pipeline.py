from pathlib import Path
import json

import pandas as pd
from openpyxl import load_workbook

from integra_pose.logic.batch_figures import export_batch_figure_bundle, render_video_quicklook_bundle
from integra_pose.utils.batch_exporter import (
    collect_batch_module_tables,
    export_batch_module_tables,
    write_batch_workbook,
)


def test_export_batch_figure_bundle_creates_manifest_and_outputs(tmp_path):
    existing_plot = tmp_path / "existing_dashboard.png"
    existing_plot.write_bytes(b"placeholder")
    manifest_path = tmp_path / "run_manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "outputs": {
                    "roi_metrics_files": {
                        "analytics_dashboard": str(existing_plot),
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    precomputed_quicklook = tmp_path / "precomputed_quicklook.png"
    precomputed_quicklook.write_bytes(b"placeholder")
    precomputed_records = [
        {
            "scope": "individual",
            "figure_type": "video_quicklook",
            "metric": "",
            "video_id": "vid_001",
            "video_name": "Mouse A Day 1",
            "group": "Control",
            "subject_id": "A1",
            "time_point": "1",
            "group_a": "",
            "group_b": "",
            "factor": "",
            "format": "png",
            "title": "Mouse A Day 1 quicklook",
            "source": "batch_quicklook",
            "path": str(precomputed_quicklook.resolve()),
        }
    ]

    video_results = [
        {
            "video_id": "vid_001",
            "video_name": "Mouse A Day 1",
            "group": "Control",
            "subject_id": "A1",
            "time_point": "1",
            "run_manifest_json": str(manifest_path),
        }
    ]
    video_summary_df = pd.DataFrame(
        [
            {
                "video_id": "vid_001",
                "video_name": "Mouse A Day 1",
                "video_path": "C:/videos/a1_d1.mp4",
                "group": "Control",
                "subject_id": "A1",
                "time_point": "1",
                "bout_count": 10,
                "roi_dwell_total_s": 20.0,
                "mean_speed_px_per_frame": 2.5,
            },
            {
                "video_id": "vid_002",
                "video_name": "Mouse A Day 2",
                "video_path": "C:/videos/a1_d2.mp4",
                "group": "Control",
                "subject_id": "A1",
                "time_point": "2",
                "bout_count": 12,
                "roi_dwell_total_s": 24.0,
                "mean_speed_px_per_frame": 2.9,
            },
            {
                "video_id": "vid_003",
                "video_name": "Mouse B Day 1",
                "video_path": "C:/videos/b1_d1.mp4",
                "group": "Treatment",
                "subject_id": "B1",
                "time_point": "1",
                "bout_count": 18,
                "roi_dwell_total_s": 31.0,
                "mean_speed_px_per_frame": 3.8,
            },
            {
                "video_id": "vid_004",
                "video_name": "Mouse B Day 2",
                "video_path": "C:/videos/b1_d2.mp4",
                "group": "Treatment",
                "subject_id": "B1",
                "time_point": "2",
                "bout_count": 21,
                "roi_dwell_total_s": 35.0,
                "mean_speed_px_per_frame": 4.1,
            },
        ]
    )
    omnibus_df = pd.DataFrame(
        [
            {"factor": "group", "metric": "bout_count", "p_adj": 0.012, "epsilon_squared": 0.42},
            {"factor": "group", "metric": "roi_dwell_total_s", "p_adj": 0.031, "epsilon_squared": 0.35},
        ]
    )
    pairwise_df = pd.DataFrame(
        [
            {
                "factor": "group",
                "metric": "bout_count",
                "group_a": "Control",
                "group_b": "Treatment",
                "cliffs_delta": -0.72,
                "p_adj": 0.011,
            },
            {
                "factor": "group",
                "metric": "roi_dwell_total_s",
                "group_a": "Control",
                "group_b": "Treatment",
                "cliffs_delta": -0.61,
                "p_adj": 0.028,
            },
        ]
    )
    module_tables = {
        "temporal_trends__duration": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Behavior": "Explore", "Time Bin Start (s)": 0, "Total_Duration_s": 11.0},
                {"video_id": "vid_002", "video_name": "Mouse A Day 2", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "2", "Behavior": "Explore", "Time Bin Start (s)": 60, "Total_Duration_s": 13.0},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Behavior": "Explore", "Time Bin Start (s)": 0, "Total_Duration_s": 19.0},
                {"video_id": "vid_004", "video_name": "Mouse B Day 2", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "2", "Behavior": "Explore", "Time Bin Start (s)": 60, "Total_Duration_s": 22.0},
            ]
        ),
        "activity_budgets__global": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Behavior": "Explore", "Proportion_of_Session": 0.42},
                {"video_id": "vid_002", "video_name": "Mouse A Day 2", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "2", "Behavior": "Explore", "Proportion_of_Session": 0.45},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Behavior": "Explore", "Proportion_of_Session": 0.58},
                {"video_id": "vid_004", "video_name": "Mouse B Day 2", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "2", "Behavior": "Explore", "Proportion_of_Session": 0.61},
            ]
        ),
        "behavior_transitions__global_matrix": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "From Behavior": "Explore", "To Behavior": "Groom", "Transition Percentage": 40.0},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "From Behavior": "Explore", "To Behavior": "Groom", "Transition Percentage": 55.0},
            ]
        ),
        "preference_indices__zone_pairwise": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Target A": "Center", "Target B": "Periphery", "Time Preference Index": -0.18},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Target A": "Center", "Target B": "Periphery", "Time Preference Index": 0.24},
            ]
        ),
        "latency_metrics__zone_latency": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Target Name": "Center", "First Entry Latency (s)": 12.0, "Second Entry Latency (s)": 44.0},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Target Name": "Center", "First Entry Latency (s)": 7.0, "Second Entry Latency (s)": 23.0},
            ]
        ),
        "normalization_summary__zone_summary": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "ROI Name": "Center", "Time per Minute (s/min)": 12.0, "Entries per Minute": 1.6, "Percent of Session": 0.22},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "ROI Name": "Center", "Time per Minute (s/min)": 18.0, "Entries per Minute": 2.1, "Percent of Session": 0.31},
            ]
        ),
        "visit_structure__zone_summary": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "ROI Name": "Center", "Visits": 5, "Mean_Duration_s": 4.2, "Median_Duration_s": 3.9},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "ROI Name": "Center", "Visits": 7, "Mean_Duration_s": 5.6, "Median_Duration_s": 5.1},
            ]
        ),
        "inter_bout_intervals__summary": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Behavior": "Explore", "Median Inter-Bout Interval (s)": 9.5, "Mean Inter-Bout Interval (s)": 10.3, "Latency to First Bout (s)": 3.0},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Behavior": "Explore", "Median Inter-Bout Interval (s)": 6.1, "Mean Inter-Bout Interval (s)": 7.0, "Latency to First Bout (s)": 2.0},
            ]
        ),
        "object_transition_analysis__matrix": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "From Object": "Object A", "To Object": "Object B", "Transition_Count": 4},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "From Object": "Object A", "To Object": "Object B", "Transition_Count": 7},
            ]
        ),
        "event_aligned_windows__summary": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": -1.0, "Behavior": "Explore", "Behavior Fraction": 0.35},
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": 0.0, "Behavior": "Explore", "Behavior Fraction": 0.62},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": -1.0, "Behavior": "Explore", "Behavior Fraction": 0.28},
                {"video_id": "vid_003", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": 0.0, "Behavior": "Explore", "Behavior Fraction": 0.54},
            ]
        ),
    }

    manifest_df, artifacts = export_batch_figure_bundle(
        video_summary_df=video_summary_df,
        omnibus_df=omnibus_df,
        pairwise_df=pairwise_df,
        output_dir=tmp_path / "figures",
        video_results=video_results,
        module_tables=module_tables,
        precomputed_figure_records=precomputed_records,
        assay_preset_key="t_maze",
        export_mode="full_bundle",
        group_col="group",
        time_col="time_point",
    )

    assert Path(artifacts.manifest_csv).is_file()
    assert Path(artifacts.assay_manifest_csv).is_file()
    assert artifacts.figure_count == len(manifest_df)
    assert {"individual", "group", "intergroup"}.issubset(set(manifest_df["scope"]))
    assert str(existing_plot.resolve()) in set(manifest_df["path"])
    assert str(precomputed_quicklook.resolve()) in set(manifest_df["path"])
    assert manifest_df["path"].map(lambda raw: Path(raw).is_file()).all()
    assert "video_quicklook" in set(manifest_df["figure_type"])
    assert "module_temporal_trend" in set(manifest_df["figure_type"])
    assert "module_activity_budget" in set(manifest_df["figure_type"])
    assert "module_behavior_transition_heatmap" in set(manifest_df["figure_type"])
    assert "module_preference_index" in set(manifest_df["figure_type"])
    assert "module_latency_heatmap" in set(manifest_df["figure_type"])
    assert "module_normalization_heatmap" in set(manifest_df["figure_type"])
    assert "module_visit_structure_heatmap" in set(manifest_df["figure_type"])
    assert "module_inter_bout_heatmap" in set(manifest_df["figure_type"])
    assert "module_object_transition_heatmap" in set(manifest_df["figure_type"])
    assert "module_event_aligned_heatmap" in set(manifest_df["figure_type"])
    assay_manifest_df = pd.read_csv(artifacts.assay_manifest_csv)
    assert not assay_manifest_df.empty
    assert set(assay_manifest_df["assay_preset"]) == {"t_maze"}
    assert "event_aligned_windows" in set(assay_manifest_df["module_key"])
    assert assay_manifest_df["figure_rank"].iloc[0] == 1
    assert assay_manifest_df["recommended_filename"].iloc[0].startswith("t_maze_")
    assert {"primary", "secondary", "supporting"}.intersection(set(assay_manifest_df["priority_tier"]))


def test_export_batch_figure_bundle_assay_shortlist_filters_visible_manifest(tmp_path):
    summary_df = pd.DataFrame(
        [
            {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "bout_count": 10},
            {"video_id": "vid_002", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "bout_count": 14},
        ]
    )
    omnibus_df = pd.DataFrame([{"factor": "group", "metric": "bout_count", "p_adj": 0.02}])
    pairwise_df = pd.DataFrame([{"factor": "group", "metric": "bout_count", "group_a": "Control", "group_b": "Treatment", "cliffs_delta": -0.5, "p_adj": 0.03}])
    module_tables = {
        "activity_budgets__global": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Behavior": "Explore", "Proportion_of_Session": 0.45},
                {"video_id": "vid_002", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Behavior": "Explore", "Proportion_of_Session": 0.58},
            ]
        ),
        "event_aligned_windows__summary": pd.DataFrame(
            [
                {"video_id": "vid_001", "video_name": "Mouse A Day 1", "video_path": "", "group": "Control", "subject_id": "A1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": 0.0, "Behavior": "Explore", "Behavior Fraction": 0.62},
                {"video_id": "vid_002", "video_name": "Mouse B Day 1", "video_path": "", "group": "Treatment", "subject_id": "B1", "time_point": "1", "Source": "zone", "Event Type": "entry", "Target Name": "Center", "Relative Time (s)": 0.0, "Behavior": "Explore", "Behavior Fraction": 0.54},
            ]
        ),
    }

    manifest_df, artifacts = export_batch_figure_bundle(
        video_summary_df=summary_df,
        omnibus_df=omnibus_df,
        pairwise_df=pairwise_df,
        output_dir=tmp_path / "assay_only_figures",
        module_tables=module_tables,
        assay_preset_key="t_maze",
        export_mode="assay_shortlist",
        group_col="group",
        time_col="time_point",
    )

    assert Path(artifacts.manifest_csv).is_file()
    assert "module_event_aligned_heatmap" in set(manifest_df["figure_type"])
    assert "module_activity_budget" not in set(manifest_df["figure_type"])


def test_render_video_quicklook_bundle_creates_individual_figures(tmp_path):
    summary_csv = tmp_path / "mouse_a_summary.csv"
    summary_csv.write_text(
        "Class Label,Track ID,Number of Bouts,Avg Bout Duration (seconds)\nExplore,1,8,2.5\nGroom,1,3,1.2\n",
        encoding="utf-8",
    )
    roi_csv = tmp_path / "mouse_a_roi.csv"
    roi_csv.write_text(
        "ROI Name,Entries,Time in ROI (s)\nCenter,4,18.0\nPeriphery,6,24.0\n",
        encoding="utf-8",
    )
    metrics_csv = tmp_path / "mouse_a_metrics_track.csv"
    metrics_csv.write_text(
        "track_id,mean_speed_px_per_frame,turn_count\n1,2.4,14\n2,2.1,11\n",
        encoding="utf-8",
    )
    video_result = {
        "video_id": "vid_001",
        "video_name": "Mouse A Day 1",
        "group": "Control",
        "subject_id": "A1",
        "time_point": "1",
        "summary_bouts_csv": str(summary_csv),
        "roi_overview_csv": str(roi_csv),
        "metrics_summary_by_track_csv": str(metrics_csv),
    }

    records = render_video_quicklook_bundle(video_result, output_dir=tmp_path / "quicklooks")
    assert records
    assert {row["figure_type"] for row in records} == {"video_quicklook"}
    assert all(Path(row["path"]).is_file() for row in records)


def test_collect_and_export_batch_module_tables(tmp_path):
    temporal_csv = tmp_path / "mouse_a_behavior_time_by_bin.csv"
    temporal_csv.write_text(
        "Time Bin Start (s),Time Bin End (s),Behavior,Total_Duration_s\n0,60,Explore,12.0\n60,120,Explore,15.0\n",
        encoding="utf-8",
    )
    budget_csv = tmp_path / "mouse_a_activity_budget_global.csv"
    budget_csv.write_text(
        "Behavior,Total_Time_s,Proportion_of_Session\nExplore,27.0,0.45\nGroom,9.0,0.15\n",
        encoding="utf-8",
    )
    manifest_path = tmp_path / "run_manifest.json"
    manifest_path.write_text(
        json.dumps(
            {
                "outputs": {
                    "modules": {
                        "temporal_trends": {
                            "files": {
                                "duration": str(temporal_csv),
                            }
                        },
                        "activity_budgets": {
                            "files": {
                                "global": str(budget_csv),
                            }
                        },
                    }
                }
            }
        ),
        encoding="utf-8",
    )
    video_results = [
        {
            "video_id": "vid_001",
            "video_name": "Mouse A Day 1",
            "video_path": "C:/videos/a1_d1.mp4",
            "group": "Control",
            "subject_id": "A1",
            "time_point": "1",
            "run_manifest_json": str(manifest_path),
        }
    ]

    bundle = collect_batch_module_tables(video_results)
    assert "temporal_trends__duration" in bundle.tables
    assert "activity_budgets__global" in bundle.tables
    assert set(bundle.file_index_df["status"]) == {"loaded"}
    temporal_table = bundle.tables["temporal_trends__duration"]
    assert {"video_id", "group", "source_module", "source_file_key"}.issubset(set(temporal_table.columns))

    file_index_df, table_index_df = export_batch_module_tables(bundle, output_dir=tmp_path / "module_tables")
    assert not file_index_df.empty
    assert not table_index_df.empty
    assert table_index_df["path"].map(lambda raw: Path(raw).is_file()).all()


def test_write_batch_workbook_adds_figures_index_hyperlinks(tmp_path):
    figure_path = tmp_path / "figures" / "summary_plot.png"
    figure_path.parent.mkdir(parents=True, exist_ok=True)
    figure_path.write_bytes(b"placeholder")
    figure_manifest_df = pd.DataFrame(
        [
            {
                "scope": "group",
                "figure_type": "group_distribution",
                "metric": "bout_count",
                "video_id": "",
                "video_name": "",
                "group": "",
                "subject_id": "",
                "time_point": "",
                "group_a": "",
                "group_b": "",
                "factor": "group",
                "format": "png",
                "title": "Bout Count by Group",
                "source": "video_summary_df",
                "path": str(figure_path.resolve()),
            }
        ]
    )
    module_file_index_df = pd.DataFrame(
        [
            {
                "video_id": "vid_001",
                "video_name": "Mouse A Day 1",
                "video_path": "C:/videos/a1_d1.mp4",
                "group": "Control",
                "subject_id": "A1",
                "time_point": "1",
                "module_key": "temporal_trends",
                "file_key": "duration",
                "table_key": "temporal_trends__duration",
                "status": "loaded",
                "row_count": 2,
                "column_count": 4,
                "path": str(figure_path.resolve()),
            }
        ]
    )
    module_table_index_df = pd.DataFrame(
        [
            {
                "table_key": "temporal_trends__duration",
                "module_key": "temporal_trends",
                "file_key": "duration",
                "row_count": 2,
                "column_count": 10,
                "video_count": 1,
                "path": str(figure_path.resolve()),
            }
        ]
    )
    assay_figure_manifest_df = pd.DataFrame(
        [
            {
                "assay_preset": "t_maze",
                "assay_label": "T-maze",
                "figure_rank": 1,
                "priority_tier": "primary",
                "selection_reason": "preset_module",
                "recommended_label": "01. Event-aligned behavior: Zone entry | Center",
                "recommended_filename": "t_maze_01_event_aligned_behavior_zone_entry_center.png",
                "module_key": "event_aligned_windows",
                "scope": "group",
                "figure_type": "module_event_aligned_heatmap",
                "metric": "Behavior Fraction",
                "video_id": "",
                "video_name": "",
                "group": "",
                "subject_id": "",
                "time_point": "",
                "group_a": "",
                "group_b": "",
                "factor": "event_aligned_windows__summary",
                "format": "png",
                "title": "Event-aligned behavior: Zone entry | Center",
                "source": "module_tables.event_aligned_windows__summary",
                "path": str(figure_path.resolve()),
            }
        ]
    )
    workbook_path = write_batch_workbook(
        workbook_path=tmp_path / "batch_results.xlsx",
        keypoint_df=pd.DataFrame(),
        kinematic_df=pd.DataFrame(),
        bout_df=pd.DataFrame(),
        video_summary_df=pd.DataFrame(),
        omnibus_df=pd.DataFrame(),
        pairwise_df=pd.DataFrame(),
        kpss_df=pd.DataFrame(),
        effects_df=pd.DataFrame(),
        analysis_coverage_df=pd.DataFrame(),
        figure_manifest_df=figure_manifest_df,
        assay_figure_manifest_df=assay_figure_manifest_df,
        module_file_index_df=module_file_index_df,
        module_table_index_df=module_table_index_df,
    )

    workbook = load_workbook(workbook_path)
    assert "Figures_Index" in workbook.sheetnames
    assert "Assay_Figure_Index" in workbook.sheetnames
    assert "Module_File_Index" in workbook.sheetnames
    assert "Module_Table_Index" in workbook.sheetnames
    sheet = workbook["Figures_Index"]
    headers = {sheet.cell(row=1, column=idx).value: idx for idx in range(1, sheet.max_column + 1)}
    path_cell = sheet.cell(row=2, column=headers["path"])
    assert path_cell.value == str(figure_path.resolve())
    assert path_cell.hyperlink is not None
    assert path_cell.hyperlink.target == str(figure_path.resolve())
    assay_sheet = workbook["Assay_Figure_Index"]
    assay_headers = {assay_sheet.cell(row=1, column=idx).value: idx for idx in range(1, assay_sheet.max_column + 1)}
    assay_path_cell = assay_sheet.cell(row=2, column=assay_headers["path"])
    assert assay_path_cell.value == str(figure_path.resolve())
    assert assay_path_cell.hyperlink is not None
    assert assay_path_cell.hyperlink.target == str(figure_path.resolve())
