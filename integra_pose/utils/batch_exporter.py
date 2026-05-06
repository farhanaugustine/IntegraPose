"""Batch-level export helpers (XLSX + consolidated CSVs)."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter


def _safe_read_csv(path: str | Path) -> pd.DataFrame:
    target = Path(str(path or "")).expanduser()
    if not target.is_file():
        return pd.DataFrame()
    try:
        # Pin UTF-8 to round-trip with the writers in this codebase
        # (which emit UTF-8) and to avoid Windows CP1252 corruption on
        # files containing accented filenames or non-ASCII metadata.
        return pd.read_csv(target, encoding="utf-8")
    except UnicodeDecodeError:
        # Best-effort recovery for legacy CSVs written before the UTF-8
        # standard was applied — try the platform default and emit no error.
        try:
            return pd.read_csv(target)
        except Exception:
            return pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _safe_read_json(path: str | Path) -> dict[str, Any]:
    target = Path(str(path or "")).expanduser()
    if not target.is_file():
        return {}
    try:
        data = json.loads(target.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _flatten_paths(payload: Any) -> list[str]:
    values: list[str] = []

    def _visit(candidate: Any) -> None:
        if candidate is None:
            return
        if isinstance(candidate, Path):
            text = str(candidate).strip()
            if text:
                values.append(text)
            return
        if isinstance(candidate, str):
            text = candidate.strip()
            if text:
                values.append(text)
            return
        if isinstance(candidate, dict):
            for value in candidate.values():
                _visit(value)
            return
        if isinstance(candidate, (list, tuple, set)):
            for value in candidate:
                _visit(value)

    _visit(payload)
    deduped: list[str] = []
    seen: set[str] = set()
    for path in values:
        key = path.lower()
        if key in seen:
            continue
        seen.add(key)
        deduped.append(path)
    return deduped


def _existing_file_count(paths: list[str]) -> int:
    count = 0
    for path in paths:
        try:
            if Path(path).expanduser().is_file():
                count += 1
        except Exception:
            continue
    return count


_ANALYTICS_MODULE_SPECS: list[tuple[str, str, str]] = [
    ("preference_indices", "Preference indices", "Pairwise preference, discrimination, and first-choice summaries"),
    ("latency_metrics", "Latency metrics", "First entry, interaction, contact, and revisit latencies"),
    ("visit_structure", "Visit structure", "Visit counts, dwell distributions, and visit-duration summaries"),
    ("object_transition_analysis", "Object transitions", "Stimulus visit sequences and object-to-object transitions"),
    ("event_aligned_windows", "Event-aligned windows", "Behavior around ROI/object entry and exit events"),
    ("normalization_summary", "Normalized summaries", "Per-minute and percent-of-session normalization tables"),
    ("behavior_transitions", "Behavior transitions", "Behavior, next behavior, transition count"),
    ("temporal_trends", "Temporal trends", "Behavior, time bin, counts, dwell duration"),
    ("activity_budgets", "Activity budgets", "Behavior, track, ROI, time allocation"),
    ("multi_animal_descriptors", "Interaction descriptors", "Track pairs, proximity, co-occurrence"),
    ("roi_time_heatmap", "ROI time heatmap", "ROI occupancy by time bin"),
    ("kinematic_descriptors", "Kinematic descriptors", "Speed, acceleration, path length"),
    ("detection_quality", "Detection quality", "Keypoint confidence, completeness, bbox stability"),
    ("inter_bout_intervals", "Inter-bout intervals", "Inter-bout interval, first-bout latency"),
    ("roi_context_windows", "ROI context windows", "Pre/post bout ROI state windows"),
    ("bout_timeline_export", "Bout timeline export", "Bout start/end and normalized timeline"),
]

MODULE_FILE_INDEX_COLUMNS = [
    "video_id",
    "video_name",
    "video_path",
    "group",
    "subject_id",
    "time_point",
    "module_key",
    "file_key",
    "table_key",
    "status",
    "row_count",
    "column_count",
    "path",
]

MODULE_TABLE_INDEX_COLUMNS = [
    "table_key",
    "module_key",
    "file_key",
    "row_count",
    "column_count",
    "video_count",
    "path",
]


@dataclass(slots=True)
class BatchModuleTableBundle:
    file_index_df: pd.DataFrame
    tables: dict[str, pd.DataFrame]


def _safe_sheet_name(name: str, *, fallback: str = "Sheet") -> str:
    cleaned = str(name or "").strip().replace("/", "_").replace("\\", "_").replace(":", "_")
    cleaned = cleaned[:31].strip()
    return cleaned or fallback


def _infer_table_key(module_key: str, file_key: str) -> str:
    return f"{str(module_key).strip()}__{str(file_key).strip()}"


def _normalize_table_export_name(table_key: str) -> str:
    safe = "".join(ch if ch.isalnum() or ch in {"_", "-"} else "_" for ch in str(table_key or "").strip())
    return safe.lower().strip("_") or "module_table"


def _module_context_from_result(result: dict[str, Any]) -> dict[str, str]:
    return {
        "video_id": str(result.get("video_id", "")).strip(),
        "video_name": str(result.get("video_name", "")).strip(),
        "video_path": str(result.get("video_path", "")).strip(),
        "group": str(result.get("group", "")).strip(),
        "subject_id": str(result.get("subject_id", "")).strip(),
        "time_point": str(result.get("time_point", "")).strip(),
    }


def _attach_module_context(
    df: pd.DataFrame,
    *,
    context: dict[str, str],
    module_key: str,
    file_key: str,
    path: str,
) -> pd.DataFrame:
    out = df.copy()
    insert_pairs = [
        ("video_id", context.get("video_id", "")),
        ("video_name", context.get("video_name", "")),
        ("video_path", context.get("video_path", "")),
        ("group", context.get("group", "")),
        ("subject_id", context.get("subject_id", "")),
        ("time_point", context.get("time_point", "")),
        ("source_module", str(module_key)),
        ("source_file_key", str(file_key)),
        ("source_file_path", str(path)),
    ]
    for offset, (column, value) in enumerate(insert_pairs):
        if column in out.columns:
            out[column] = value
        else:
            out.insert(offset, column, value)
    return out


def collect_batch_module_tables(video_results: list[dict[str, Any]]) -> BatchModuleTableBundle:
    file_rows: list[dict[str, Any]] = []
    table_chunks: dict[str, list[pd.DataFrame]] = {}

    for result in video_results:
        context = _module_context_from_result(result)
        manifest = _safe_read_json(result.get("run_manifest_json", ""))
        outputs = manifest.get("outputs", {}) if isinstance(manifest.get("outputs"), dict) else {}
        module_outputs = outputs.get("modules", {}) if isinstance(outputs.get("modules"), dict) else {}
        for module_key, payload in module_outputs.items():
            payload_dict = payload if isinstance(payload, dict) else {}
            files = payload_dict.get("files") if isinstance(payload_dict.get("files"), dict) else {}
            for file_key, raw_path in files.items():
                path_text = str(raw_path or "").strip()
                table_key = _infer_table_key(str(module_key), str(file_key))
                resolved_path = Path(path_text).expanduser()
                row = dict(context)
                row.update(
                    {
                        "module_key": str(module_key),
                        "file_key": str(file_key),
                        "table_key": table_key,
                        "status": "missing",
                        "row_count": 0,
                        "column_count": 0,
                        "path": path_text,
                    }
                )
                if not path_text:
                    file_rows.append(row)
                    continue
                if not resolved_path.is_file():
                    file_rows.append(row)
                    continue
                if resolved_path.suffix.lower() != ".csv":
                    row["status"] = "non_tabular"
                    file_rows.append(row)
                    continue
                df = _safe_read_csv(resolved_path)
                row["status"] = "loaded"
                row["row_count"] = int(len(df.index))
                row["column_count"] = int(len(df.columns))
                row["path"] = str(resolved_path.resolve())
                file_rows.append(row)
                if not df.empty or len(df.columns) > 0:
                    enriched = _attach_module_context(
                        df,
                        context=context,
                        module_key=str(module_key),
                        file_key=str(file_key),
                        path=str(resolved_path.resolve()),
                    )
                    table_chunks.setdefault(table_key, []).append(enriched)

    file_index_df = (
        pd.DataFrame(file_rows, columns=MODULE_FILE_INDEX_COLUMNS)
        if file_rows
        else pd.DataFrame(columns=MODULE_FILE_INDEX_COLUMNS)
    )
    tables = {
        table_key: pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
        for table_key, chunks in table_chunks.items()
    }
    return BatchModuleTableBundle(file_index_df=file_index_df, tables=tables)


def export_batch_module_tables(
    bundle: BatchModuleTableBundle,
    *,
    output_dir: str | Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    output_root = Path(output_dir).expanduser().resolve()
    output_root.mkdir(parents=True, exist_ok=True)

    file_index_df = bundle.file_index_df if bundle.file_index_df is not None else pd.DataFrame(columns=MODULE_FILE_INDEX_COLUMNS)
    file_index_csv = output_root / "module_file_index.csv"
    file_index_df.to_csv(file_index_csv, index=False)

    table_rows: list[dict[str, Any]] = []
    for table_key, df in sorted(bundle.tables.items()):
        export_name = _normalize_table_export_name(table_key)
        csv_path = output_root / f"{export_name}.csv"
        (df if df is not None else pd.DataFrame()).to_csv(csv_path, index=False)
        module_key, _, file_key = table_key.partition("__")
        video_count = int(df["video_id"].nunique()) if df is not None and not df.empty and "video_id" in df.columns else 0
        table_rows.append(
            {
                "table_key": table_key,
                "module_key": module_key,
                "file_key": file_key,
                "row_count": int(len(df.index)) if df is not None else 0,
                "column_count": int(len(df.columns)) if df is not None else 0,
                "video_count": video_count,
                "path": str(csv_path),
            }
        )

    table_index_df = (
        pd.DataFrame(table_rows, columns=MODULE_TABLE_INDEX_COLUMNS)
        if table_rows
        else pd.DataFrame(columns=MODULE_TABLE_INDEX_COLUMNS)
    )
    table_index_csv = output_root / "module_table_index.csv"
    table_index_df.to_csv(table_index_csv, index=False)
    return file_index_df, table_index_df


def build_analysis_coverage(video_results: list[dict[str, Any]]) -> pd.DataFrame:
    columns = [
        "video_id",
        "video_name",
        "group",
        "subject_id",
        "time_point",
        "analysis_key",
        "analysis_label",
        "enabled",
        "status",
        "variables",
        "outputs_expected",
        "outputs_found",
        "completion_ratio",
        "output_files",
    ]
    rows: list[dict[str, Any]] = []

    for result in video_results:
        manifest = _safe_read_json(result.get("run_manifest_json", ""))
        parameters = manifest.get("parameters", {}) if isinstance(manifest.get("parameters"), dict) else {}
        outputs = manifest.get("outputs", {}) if isinstance(manifest.get("outputs"), dict) else {}
        module_outputs = outputs.get("modules", {}) if isinstance(outputs.get("modules"), dict) else {}
        enabled_modules = {
            str(name).strip()
            for name in (parameters.get("enabled_modules", []) if isinstance(parameters.get("enabled_modules"), list) else [])
            if str(name).strip()
        }

        video_context = {
            "video_id": str(result.get("video_id", "")).strip(),
            "video_name": str(result.get("video_name", "")).strip(),
            "group": str(result.get("group", "")).strip(),
            "subject_id": str(result.get("subject_id", "")).strip(),
            "time_point": str(result.get("time_point", "")).strip(),
        }

        def _add_row(
            *,
            key: str,
            label: str,
            enabled: bool,
            variables: str,
            paths_payload: Any,
            required_outputs_min: int = 1,
        ) -> None:
            file_paths = _flatten_paths(paths_payload)
            outputs_expected = int(len(file_paths))
            outputs_found = _existing_file_count(file_paths)
            if not enabled:
                status = "not_enabled"
                completion_ratio = 0.0
            elif outputs_expected <= 0:
                status = "enabled_no_output"
                completion_ratio = 0.0
            elif outputs_found >= max(1, int(required_outputs_min or 1)) and outputs_found < outputs_expected:
                status = "partial"
                completion_ratio = float(outputs_found / outputs_expected)
            elif outputs_found >= outputs_expected:
                status = "completed"
                completion_ratio = 1.0
            else:
                status = "enabled_no_output"
                completion_ratio = 0.0

            row = dict(video_context)
            row.update(
                {
                    "analysis_key": key,
                    "analysis_label": label,
                    "enabled": bool(enabled),
                    "status": status,
                    "variables": variables,
                    "outputs_expected": outputs_expected,
                    "outputs_found": int(outputs_found),
                    "completion_ratio": completion_ratio,
                    "output_files": "; ".join(file_paths),
                }
            )
            rows.append(row)

        _add_row(
            key="core_bouts",
            label="Core bout analysis",
            enabled=True,
            variables="Track ID, behavior, start/end frame, bout duration",
            paths_payload=[
                outputs.get("detailed_bouts_csv"),
                outputs.get("summary_csv"),
                outputs.get("excel_summary"),
                result.get("detailed_bouts_csv"),
                result.get("summary_bouts_csv"),
            ],
            required_outputs_min=2,
        )

        roi_files = _flatten_paths(
            [
                outputs.get("roi_metrics_files"),
                result.get("roi_overview_csv"),
            ]
        )
        roi_enabled = bool(parameters.get("use_rois", bool(roi_files)))
        _add_row(
            key="roi_metrics",
            label="ROI metrics",
            enabled=roi_enabled,
            variables="ROI entries/exits, dwell time, transitions, ROI behavior counts",
            paths_payload=roi_files,
            required_outputs_min=1,
        )

        object_files = _flatten_paths(
            [
                outputs.get("object_interaction_files"),
                result.get("object_interactions_csv"),
            ]
        )
        object_enabled = bool(parameters.get("object_interaction_enabled", bool(object_files)))
        _add_row(
            key="object_interactions",
            label="Object interactions",
            enabled=object_enabled,
            variables="Object ROI entries/exits, interaction time, per-frame memberships",
            paths_payload=object_files,
            required_outputs_min=1,
        )

        for module_key, module_label, variables in _ANALYTICS_MODULE_SPECS:
            payload = module_outputs.get(module_key)
            payload_dict = payload if isinstance(payload, dict) else {}
            files = payload_dict.get("files") if isinstance(payload_dict.get("files"), dict) else {}
            if enabled_modules:
                enabled = module_key in enabled_modules
            else:
                enabled = bool(payload_dict)
            _add_row(
                key=module_key,
                label=module_label,
                enabled=enabled,
                variables=variables,
                paths_payload=files,
                required_outputs_min=1,
            )

    if not rows:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(rows, columns=columns)


def _attach_video_context(
    df: pd.DataFrame,
    *,
    video_id: str,
    video_name: str,
    video_path: str,
    group: str,
    subject_id: str = "",
    time_point: str = "",
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out = df.copy()
    out.insert(0, "video_id", video_id)
    out.insert(1, "video_name", video_name)
    out.insert(2, "video_path", video_path)
    out.insert(3, "group", group)
    out.insert(4, "subject_id", subject_id)
    out.insert(5, "time_point", time_point)
    return out


def check_fps_consistency(
    video_results: list[dict[str, Any]],
    *,
    tolerance: float = 0.5,
    log_fn: Any = None,
) -> dict[str, Any]:
    """Read each video's ``inference_metadata.json`` and warn on FPS heterogeneity.

    Cross-video aggregation (e.g. concatenating "events per second" across
    videos in ``collect_batch_frames``) is only meaningful when every video
    shares a common frame rate. This helper does *not* refuse the aggregation
    — it surfaces the disagreement so the user can decide whether the cross-
    video figures are interpretable.

    Returns a dict ``{"by_video": {video_id: fps}, "consistent": bool,
    "min_fps": float, "max_fps": float, "spread": float, "n_videos": int}``.
    """
    from pathlib import Path as _Path

    by_video: dict[str, float] = {}
    sources: dict[str, str] = {}
    missing: list[str] = []
    for result in video_results or []:
        video_id = str(result.get("video_id") or "").strip()
        if not video_id:
            continue
        run_dir = result.get("run_output_dir") or ""
        meta_path = _Path(str(run_dir)) / "inference_metadata.json" if run_dir else None
        if not (meta_path and meta_path.is_file()):
            missing.append(video_id)
            continue
        try:
            payload = _safe_read_json(meta_path)
        except Exception:
            missing.append(video_id)
            continue
        try:
            fps = float(payload.get("fps") or 0.0)
        except (TypeError, ValueError):
            fps = 0.0
        if fps > 0:
            by_video[video_id] = fps
            sources[video_id] = str(payload.get("fps_source") or "")

    fps_values = list(by_video.values())
    summary: dict[str, Any] = {
        "by_video": by_video,
        "fps_sources": sources,
        "missing": missing,
        "n_videos": len(by_video),
    }
    if not fps_values:
        summary.update({"consistent": True, "min_fps": 0.0, "max_fps": 0.0, "spread": 0.0})
        return summary
    min_fps = min(fps_values)
    max_fps = max(fps_values)
    spread = max_fps - min_fps
    consistent = spread <= float(tolerance)
    summary.update({"consistent": consistent, "min_fps": min_fps, "max_fps": max_fps, "spread": spread})

    if log_fn is not None:
        try:
            if missing:
                log_fn(
                    f"FPS consistency check: {len(missing)} video(s) lacked "
                    f"inference_metadata.json and were skipped from the cross-"
                    f"video FPS audit.",
                    "INFO",
                )
            if not consistent:
                pretty = ", ".join(
                    f"{vid}={fps:g}fps" for vid, fps in sorted(by_video.items())
                )
                log_fn(
                    (
                        f"Cross-video FPS heterogeneity detected (spread={spread:g} "
                        f"fps, range {min_fps:g}-{max_fps:g}). Per-second metrics "
                        f"aggregated across these videos may not be directly "
                        f"comparable. Per-video FPS: {pretty}."
                    ),
                    "WARNING",
                )
            else:
                log_fn(
                    f"FPS consistency check passed: all {len(by_video)} video(s) "
                    f"share FPS within {tolerance:g} (range {min_fps:g}-{max_fps:g}).",
                    "INFO",
                )
        except Exception:
            pass
    return summary


def collect_batch_frames(video_results: list[dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    keypoint_rows: list[pd.DataFrame] = []
    kinematic_rows: list[pd.DataFrame] = []
    bout_rows: list[pd.DataFrame] = []

    for result in video_results:
        video_id = str(result.get("video_id", "")).strip()
        video_name = str(result.get("video_name", "")).strip()
        video_path = str(result.get("video_path", "")).strip()
        group = str(result.get("group", "")).strip()
        subject_id = str(result.get("subject_id", "")).strip()
        time_point = str(result.get("time_point", "")).strip()

        labels_csv = _safe_read_csv(result.get("labels_csv", ""))
        if not labels_csv.empty:
            keypoint_rows.append(
                _attach_video_context(
                    labels_csv,
                    video_id=video_id,
                    video_name=video_name,
                    video_path=video_path,
                    group=group,
                    subject_id=subject_id,
                    time_point=time_point,
                )
            )

        metrics_track = _safe_read_csv(result.get("metrics_summary_by_track_csv", ""))
        if metrics_track.empty:
            metrics_track = _safe_read_csv(result.get("metrics_csv", ""))
        if not metrics_track.empty:
            kinematic_rows.append(
                _attach_video_context(
                    metrics_track,
                    video_id=video_id,
                    video_name=video_name,
                    video_path=video_path,
                    group=group,
                    subject_id=subject_id,
                    time_point=time_point,
                )
            )

        detailed = _safe_read_csv(result.get("detailed_bouts_csv", ""))
        if not detailed.empty:
            bout_rows.append(
                _attach_video_context(
                    detailed,
                    video_id=video_id,
                    video_name=video_name,
                    video_path=video_path,
                    group=group,
                    subject_id=subject_id,
                    time_point=time_point,
                )
            )

        roi_overview = _safe_read_csv(result.get("roi_overview_csv", ""))
        if not roi_overview.empty:
            roi_df = _attach_video_context(
                roi_overview,
                video_id=video_id,
                video_name=video_name,
                video_path=video_path,
                group=group,
                subject_id=subject_id,
                time_point=time_point,
            )
            roi_df.insert(4, "record_type", "roi_overview")
            bout_rows.append(roi_df)

        object_interactions = _safe_read_csv(result.get("object_interactions_csv", ""))
        if not object_interactions.empty:
            object_df = _attach_video_context(
                object_interactions,
                video_id=video_id,
                video_name=video_name,
                video_path=video_path,
                group=group,
                subject_id=subject_id,
                time_point=time_point,
            )
            object_df.insert(4, "record_type", "object_interactions")
            bout_rows.append(object_df)

    keypoint_df = pd.concat(keypoint_rows, ignore_index=True) if keypoint_rows else pd.DataFrame()
    kinematic_df = pd.concat(kinematic_rows, ignore_index=True) if kinematic_rows else pd.DataFrame()
    bout_df = pd.concat(bout_rows, ignore_index=True) if bout_rows else pd.DataFrame()
    return keypoint_df, kinematic_df, bout_df


def build_video_summary(video_results: list[dict[str, Any]]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for result in video_results:
        detailed = _safe_read_csv(result.get("detailed_bouts_csv", ""))
        roi_overview = _safe_read_csv(result.get("roi_overview_csv", ""))
        metrics_track = _safe_read_csv(result.get("metrics_summary_by_track_csv", ""))
        if metrics_track.empty:
            metrics_track = _safe_read_csv(result.get("metrics_csv", ""))
        object_interactions = _safe_read_csv(result.get("object_interactions_csv", ""))

        durations = pd.to_numeric(detailed.get("Duration (s)"), errors="coerce") if "Duration (s)" in detailed.columns else pd.Series(dtype=float)
        total_duration = float(durations.fillna(0).sum()) if not durations.empty else np.nan
        mean_duration = float(durations.dropna().mean()) if not durations.dropna().empty else np.nan
        bout_count = int(len(detailed)) if not detailed.empty else 0

        entries_total = np.nan
        exits_total = np.nan
        dwell_total_s = np.nan
        if not roi_overview.empty:
            if "Entries" in roi_overview.columns:
                entries_total = float(pd.to_numeric(roi_overview["Entries"], errors="coerce").fillna(0).sum())
            if "Exits" in roi_overview.columns:
                exits_total = float(pd.to_numeric(roi_overview["Exits"], errors="coerce").fillna(0).sum())
            if "Time in ROI (s)" in roi_overview.columns:
                dwell_total_s = float(pd.to_numeric(roi_overview["Time in ROI (s)"], errors="coerce").fillna(0).sum())

        mean_speed = np.nan
        mean_turns = np.nan
        if not metrics_track.empty:
            if "mean_speed_px_per_frame" in metrics_track.columns:
                mean_speed = float(pd.to_numeric(metrics_track["mean_speed_px_per_frame"], errors="coerce").dropna().mean())
            if "turn_count" in metrics_track.columns:
                mean_turns = float(pd.to_numeric(metrics_track["turn_count"], errors="coerce").dropna().mean())

        object_interaction_time_s = np.nan
        object_interaction_entries = np.nan
        if not object_interactions.empty:
            if "Time Interacting (s)" in object_interactions.columns:
                object_interaction_time_s = float(
                    pd.to_numeric(object_interactions["Time Interacting (s)"], errors="coerce").fillna(0).sum()
                )
            if "Entries" in object_interactions.columns:
                object_interaction_entries = float(
                    pd.to_numeric(object_interactions["Entries"], errors="coerce").fillna(0).sum()
                )

        rows.append(
            {
                "video_id": str(result.get("video_id", "")).strip(),
                "video_name": str(result.get("video_name", "")).strip(),
                "video_path": str(result.get("video_path", "")).strip(),
                "group": str(result.get("group", "")).strip(),
                "subject_id": str(result.get("subject_id", "")).strip(),
                "time_point": str(result.get("time_point", "")).strip(),
                "bout_count": bout_count,
                "bout_duration_total_s": total_duration,
                "bout_duration_mean_s": mean_duration,
                "roi_entries_total": entries_total,
                "roi_exits_total": exits_total,
                "roi_dwell_total_s": dwell_total_s,
                "mean_speed_px_per_frame": mean_speed,
                "mean_turn_count": mean_turns,
                "object_interaction_total_s": object_interaction_time_s,
                "object_interaction_entries_total": object_interaction_entries,
            }
        )

    return pd.DataFrame(rows)


def write_batch_workbook(
    *,
    workbook_path: str | Path,
    keypoint_df: pd.DataFrame,
    kinematic_df: pd.DataFrame,
    bout_df: pd.DataFrame,
    video_summary_df: pd.DataFrame,
    omnibus_df: pd.DataFrame,
    pairwise_df: pd.DataFrame,
    kpss_df: pd.DataFrame,
    effects_df: pd.DataFrame,
    analysis_coverage_df: pd.DataFrame | None = None,
    figure_manifest_df: pd.DataFrame | None = None,
    assay_figure_manifest_df: pd.DataFrame | None = None,
    module_file_index_df: pd.DataFrame | None = None,
    module_table_index_df: pd.DataFrame | None = None,
) -> str:
    target = Path(workbook_path).expanduser().resolve()
    target.parent.mkdir(parents=True, exist_ok=True)

    def _autofit_and_link(sheet_name: str, *, hyperlink_columns: set[str] | None = None) -> None:
        worksheet = writer.sheets.get(sheet_name)
        if worksheet is None or worksheet.max_row <= 0 or worksheet.max_column <= 0:
            return
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
        hyperlink_targets = set(hyperlink_columns or set())
        for column_index, header in enumerate(headers, start=1):
            column_letter = get_column_letter(column_index)
            values = [str(header or "")]
            for row_index in range(2, min(worksheet.max_row, 200) + 1):
                cell_value = worksheet.cell(row=row_index, column=column_index).value
                values.append(str(cell_value or ""))
                if header in hyperlink_targets and cell_value:
                    worksheet.cell(row=row_index, column=column_index).hyperlink = str(cell_value)
                    worksheet.cell(row=row_index, column=column_index).style = "Hyperlink"
            worksheet.column_dimensions[column_letter].width = min(72, max(12, max(len(value) for value in values) + 2))

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        (keypoint_df if not keypoint_df.empty else pd.DataFrame(columns=["video_id", "video_name"])).to_excel(
            writer, index=False, sheet_name="Keypoint_Estimation"
        )
        _autofit_and_link("Keypoint_Estimation")
        (kinematic_df if not kinematic_df.empty else pd.DataFrame(columns=["video_id", "video_name"])).to_excel(
            writer, index=False, sheet_name="Kinematic_Outputs"
        )
        _autofit_and_link("Kinematic_Outputs")
        (bout_df if not bout_df.empty else pd.DataFrame(columns=["video_id", "video_name"])).to_excel(
            writer, index=False, sheet_name="Bout_ROI_Metrics"
        )
        _autofit_and_link("Bout_ROI_Metrics")
        (video_summary_df if not video_summary_df.empty else pd.DataFrame(columns=["video_id", "video_name"])).to_excel(
            writer, index=False, sheet_name="Group_Video_Summary"
        )
        _autofit_and_link("Group_Video_Summary")
        (omnibus_df if not omnibus_df.empty else pd.DataFrame(columns=["metric"])).to_excel(
            writer, index=False, sheet_name="Stats_Omnibus"
        )
        _autofit_and_link("Stats_Omnibus")
        (pairwise_df if not pairwise_df.empty else pd.DataFrame(columns=["metric"])).to_excel(
            writer, index=False, sheet_name="Stats_Pairwise"
        )
        _autofit_and_link("Stats_Pairwise")
        (kpss_df if not kpss_df.empty else pd.DataFrame(columns=["metric"])).to_excel(
            writer, index=False, sheet_name="Stats_KPSS"
        )
        _autofit_and_link("Stats_KPSS")
        (effects_df if not effects_df.empty else pd.DataFrame(columns=["metric"])).to_excel(
            writer, index=False, sheet_name="Stats_Effect_Sizes"
        )
        _autofit_and_link("Stats_Effect_Sizes")
        coverage_df = analysis_coverage_df if analysis_coverage_df is not None else pd.DataFrame()
        if coverage_df.empty:
            coverage_df = pd.DataFrame(
                columns=[
                    "video_id",
                    "video_name",
                    "group",
                    "subject_id",
                    "time_point",
                    "analysis_key",
                    "analysis_label",
                    "enabled",
                    "status",
                    "variables",
                    "outputs_expected",
                    "outputs_found",
                    "completion_ratio",
                    "output_files",
                ]
            )
        coverage_df.to_excel(writer, index=False, sheet_name="Analysis_Coverage")
        _autofit_and_link("Analysis_Coverage")
        module_files_df = module_file_index_df if module_file_index_df is not None else pd.DataFrame()
        if module_files_df.empty:
            module_files_df = pd.DataFrame(columns=MODULE_FILE_INDEX_COLUMNS)
        module_files_df.to_excel(writer, index=False, sheet_name="Module_File_Index")
        _autofit_and_link("Module_File_Index", hyperlink_columns={"path"})
        module_tables_df = module_table_index_df if module_table_index_df is not None else pd.DataFrame()
        if module_tables_df.empty:
            module_tables_df = pd.DataFrame(columns=MODULE_TABLE_INDEX_COLUMNS)
        module_tables_df.to_excel(writer, index=False, sheet_name="Module_Table_Index")
        _autofit_and_link("Module_Table_Index", hyperlink_columns={"path"})
        figures_df = figure_manifest_df if figure_manifest_df is not None else pd.DataFrame()
        if figures_df.empty:
            figures_df = pd.DataFrame(
                columns=[
                    "scope",
                    "figure_type",
                    "metric",
                    "video_id",
                    "video_name",
                    "group",
                    "subject_id",
                    "time_point",
                    "group_a",
                    "group_b",
                    "factor",
                    "format",
                    "title",
                    "source",
                    "path",
                ]
            )
        figures_df.to_excel(writer, index=False, sheet_name="Figures_Index")
        _autofit_and_link("Figures_Index", hyperlink_columns={"path"})
        assay_figures_df = assay_figure_manifest_df if assay_figure_manifest_df is not None else pd.DataFrame()
        if assay_figures_df.empty:
            assay_figures_df = pd.DataFrame(
                columns=[
                    "assay_preset",
                    "assay_label",
                    "figure_rank",
                    "priority_tier",
                    "selection_reason",
                    "recommended_label",
                    "recommended_filename",
                    "module_key",
                    "scope",
                    "figure_type",
                    "metric",
                    "video_id",
                    "video_name",
                    "group",
                    "subject_id",
                    "time_point",
                    "group_a",
                    "group_b",
                    "factor",
                    "format",
                    "title",
                    "source",
                    "path",
                ]
            )
        assay_figures_df.to_excel(writer, index=False, sheet_name="Assay_Figure_Index")
        _autofit_and_link("Assay_Figure_Index", hyperlink_columns={"path"})
    return str(target)
